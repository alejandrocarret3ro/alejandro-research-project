"""
Comprehensive Benchmark for Blind Video Denoisers
==================================================
Compares your UNet denoiser against FastDVDNet and VRT across noise ranges.

Metrics:
- PSNR (Peak Signal-to-Noise Ratio): Measures pixel-level accuracy in dB.
  Computed as 10 * log10(MAX^2 / MSE). Higher = better. ~30dB is good, ~40dB is excellent.
  Directly reflects how well the MSE loss was minimized during training.

- SSIM (Structural Similarity Index): Measures perceptual quality (0 to 1).
  Compares luminance, contrast, and local structure between patches.
  Unlike PSNR, SSIM penalizes blurriness and structural distortion. 1.0 = identical.

How benchmarking works:
1. For each noise range (e.g., 5-25, 25-45, ...):
   - Take N frames from DAVIS validation videos
   - Add Gaussian noise with std uniformly sampled from that range
   - Run each denoiser on the same noisy input
   - Compute PSNR and SSIM of denoised output vs. clean ground truth
   - The noisy PSNR/SSIM is also recorded as a baseline (no denoising)

2. All denoisers see the exact same noisy frames (seeded RNG), so comparison is fair.

3. Results are stored as a JSON and printed as a formatted table.

Important notes on the comparison models:
- FastDVDNet was trained on noise sigma [5, 55]. It will degrade significantly
  above sigma 55 since it has never seen such noise levels during training.
- VRT was trained on fixed sigma values (e.g., sigma=30, 50). Similar limitation.
- This model was trained on sigma [5, 255], so it should handle all ranges.
  The benchmark will clearly show where each model's training range matters.
"""

import torch
import torch.nn as nn
import numpy as np
import os
import json
import time
from pathlib import Path
from PIL import Image
from skimage.metrics import peak_signal_noise_ratio as psnr
from skimage.metrics import structural_similarity as ssim
from datetime import datetime


# ============================================================================
#  FastDVDNet model definition (self-contained, no external deps needed)
# ============================================================================

class CvBlock(nn.Module):
    """Conv2D + BN + ReLU block for FastDVDNet."""
    def __init__(self, in_ch, out_ch):
        super().__init__()
        self.block = nn.Sequential(
            nn.Conv2d(in_ch, out_ch, 3, padding=1, bias=False),
            nn.BatchNorm2d(out_ch),
            nn.ReLU(inplace=True)
        )
    def forward(self, x):
        return self.block(x)


class InputCvBlock(nn.Module):
    """First block: takes concatenated noisy frames + noise map."""
    def __init__(self, num_in_frames, out_ch):
        super().__init__()
        self.block = nn.Sequential(
            nn.Conv2d(num_in_frames * (3 + 1), out_ch, 3, padding=1, bias=False),
            nn.BatchNorm2d(out_ch),
            nn.ReLU(inplace=True)
        )
    def forward(self, x):
        return self.block(x)


class DownBlock(nn.Module):
    """Downsample + 2 conv blocks."""
    def __init__(self, in_ch, out_ch):
        super().__init__()
        self.conv1 = CvBlock(in_ch, out_ch)
        self.conv2 = CvBlock(out_ch, out_ch)
        self.down = nn.MaxPool2d(2)
    def forward(self, x):
        x = self.down(x)
        x = self.conv1(x)
        return self.conv2(x)


class UpBlock(nn.Module):
    """Upsample + 2 conv blocks."""
    def __init__(self, in_ch, out_ch):
        super().__init__()
        self.up = nn.Upsample(scale_factor=2, mode='bilinear', align_corners=False)
        self.conv1 = CvBlock(in_ch, out_ch)
        self.conv2 = CvBlock(out_ch, out_ch)
    def forward(self, x, skip):
        x = self.up(x)
        x = torch.cat([x, skip], dim=1)
        x = self.conv1(x)
        return self.conv2(x)


class OutputCvBlock(nn.Module):
    """Final output block."""
    def __init__(self, in_ch, out_ch):
        super().__init__()
        self.block = nn.Conv2d(in_ch, out_ch, 3, padding=1, bias=False)
    def forward(self, x):
        return self.block(x)


class DenBlock(nn.Module):
    """Denoising UNet block used in FastDVDNet (processes 3 frames)."""
    def __init__(self, num_in_frames=3):
        super().__init__()
        ch = 32
        self.inc = InputCvBlock(num_in_frames, ch)
        self.down1 = DownBlock(ch, 2 * ch)
        self.down2 = DownBlock(2 * ch, 4 * ch)
        self.up2 = UpBlock(4 * ch + 2 * ch, 2 * ch)
        self.up1 = UpBlock(2 * ch + ch, ch)
        self.outc = OutputCvBlock(ch, 3)

    def forward(self, x):
        x1 = self.inc(x)
        x2 = self.down1(x1)
        x3 = self.down2(x2)
        x = self.up2(x3, x2)
        x = self.up1(x, x1)
        return self.outc(x)


class FastDVDNet(nn.Module):
    """
    FastDVDNet: two-stage UNet for video denoising.
    Takes 5 consecutive frames + noise sigma map.
    Stage 1: Denoise (frames 0,1,2) and (frames 2,3,4) separately.
    Stage 2: Combine both outputs with frame 2 for final denoised frame.
    """
    def __init__(self):
        super().__init__()
        self.den1 = DenBlock(num_in_frames=3)
        self.den2 = DenBlock(num_in_frames=3)

    def forward(self, x, noise_map):
        """
        Args:
            x: (B, 5, 3, H, W) — 5 consecutive frames
            noise_map: (B, 1, H, W) — noise level map (sigma/255)
        Returns:
            denoised center frame: (B, 3, H, W)
        """
        # Stage 1: process two groups of 3 frames
        # Group 1: frames 0, 1, 2
        x0 = torch.cat([x[:, 0], x[:, 1], x[:, 2]], dim=1)  # (B, 9, H, W)
        nm0 = noise_map.repeat(1, 3, 1, 1)  # (B, 3, H, W) — one per frame
        inp0 = torch.cat([x0, nm0], dim=1)  # (B, 12, H, W)
        h0 = self.den1(inp0)

        # Group 2: frames 2, 3, 4
        x1 = torch.cat([x[:, 2], x[:, 3], x[:, 4]], dim=1)
        inp1 = torch.cat([x1, nm0], dim=1)
        h1 = self.den1(inp1)

        # Stage 2: combine with center frame
        center = x[:, 2]  # (B, 3, H, W)
        x2 = torch.cat([h0, center, h1], dim=1)  # (B, 9, H, W)
        nm2 = noise_map.repeat(1, 3, 1, 1)
        inp2 = torch.cat([x2, nm2], dim=1)  # (B, 12, H, W)

        return self.den2(inp2)


# ============================================================================
#  Denoiser wrappers (unified interface for benchmarking)
# ============================================================================

class YourDenoiserWrapper:
    """Wrapper for your BlindVideoDenoiserUNet (3-frame temporal)."""

    def __init__(self, model, device='cuda'):
        self.model = model.to(device).eval()
        self.device = device
        self.name = "YourUNet"

    @torch.no_grad()
    def denoise(self, noisy_frames, noise_sigma, resize_to=(256, 256)):
        """
        Args:
            noisy_frames: list of numpy arrays (H, W, 3) float32 [0, 1]
            noise_sigma: noise std (0-255 scale, not used by model — it's blind)
            resize_to: processing resolution
        Returns:
            list of denoised numpy arrays (H, W, 3) float32 [0, 1]
        """
        denoised = []
        for i in range(len(noisy_frames)):
            prev_idx = max(0, i - 1)
            next_idx = min(len(noisy_frames) - 1, i + 1)

            # Resize to processing resolution
            prev = self._resize(noisy_frames[prev_idx], resize_to)
            curr = self._resize(noisy_frames[i], resize_to)
            nxt = self._resize(noisy_frames[next_idx], resize_to)

            # Build triplet (9, H, W)
            triplet = np.concatenate([prev, curr, nxt], axis=2)  # (H, W, 9)
            triplet_t = torch.from_numpy(triplet).permute(2, 0, 1).unsqueeze(0).float().to(self.device)

            out = self.model(triplet_t)
            out = torch.clamp(out.squeeze(0), 0, 1).cpu().permute(1, 2, 0).numpy()

            # Resize back to original
            orig_h, orig_w = noisy_frames[i].shape[:2]
            out_pil = Image.fromarray((out * 255).astype(np.uint8))
            out_pil = out_pil.resize((orig_w, orig_h), Image.LANCZOS)
            denoised.append(np.array(out_pil, dtype=np.float32) / 255.0)

        return denoised

    def _resize(self, img_np, size):
        """Resize (H, W, 3) float32 [0,1] numpy array."""
        pil = Image.fromarray((img_np * 255).astype(np.uint8))
        pil = pil.resize((size[1], size[0]), Image.BILINEAR)
        return np.array(pil, dtype=np.float32) / 255.0


class FastDVDNetWrapper:
    """Wrapper for FastDVDNet (5-frame temporal + noise map)."""

    def __init__(self, model_path, device='cuda'):
        self.device = device
        self.name = "FastDVDNet"
        self.model = FastDVDNet().to(device)

        # Load pretrained weights
        state_dict = torch.load(model_path, map_location=device)
        # Handle different checkpoint formats
        if isinstance(state_dict, dict) and 'state_dict' in state_dict:
            state_dict = state_dict['state_dict']
        self.model.load_state_dict(state_dict)
        self.model.eval()
        print(f"FastDVDNet loaded from {model_path}")

    @torch.no_grad()
    def denoise(self, noisy_frames, noise_sigma, resize_to=(256, 256)):
        """
        Args:
            noisy_frames: list of numpy arrays (H, W, 3) float32 [0, 1]
            noise_sigma: noise std (0-255 scale) — FastDVDNet uses this as input
            resize_to: processing resolution
        Returns:
            list of denoised numpy arrays (H, W, 3) float32 [0, 1]
        """
        denoised = []
        n = len(noisy_frames)

        for i in range(n):
            # Get 5 frames with boundary handling
            indices = [
                max(0, i - 2), max(0, i - 1), i,
                min(n - 1, i + 1), min(n - 1, i + 2)
            ]

            frames_resized = []
            for idx in indices:
                resized = self._resize(noisy_frames[idx], resize_to)
                frames_resized.append(resized)

            # Stack to (1, 5, 3, H, W)
            frames_t = torch.stack([
                torch.from_numpy(f).permute(2, 0, 1) for f in frames_resized
            ]).unsqueeze(0).float().to(self.device)

            # Noise map (sigma / 255 as input)
            noise_map = torch.full(
                (1, 1, resize_to[0], resize_to[1]),
                noise_sigma / 255.0,
                device=self.device
            )

            out = self.model(frames_t, noise_map)
            out = torch.clamp(out.squeeze(0), 0, 1).cpu().permute(1, 2, 0).numpy()

            # Resize back
            orig_h, orig_w = noisy_frames[i].shape[:2]
            out_pil = Image.fromarray((out * 255).astype(np.uint8))
            out_pil = out_pil.resize((orig_w, orig_h), Image.LANCZOS)
            denoised.append(np.array(out_pil, dtype=np.float32) / 255.0)

        return denoised

    def _resize(self, img_np, size):
        pil = Image.fromarray((img_np * 255).astype(np.uint8))
        pil = pil.resize((size[1], size[0]), Image.BILINEAR)
        return np.array(pil, dtype=np.float32) / 255.0


class VRTPublishedResults:
    """
    Uses VRT's published PSNR/SSIM values from the paper instead of running inference.

    VRT is too heavy to run efficiently on Colab (~35M params, requires tiling,
    custom CUDA ops, and 10+ GB VRAM). Instead, we include their published numbers
    on DAVIS for the sigma values they tested (10, 20, 30, 40, 50) and mark
    other noise ranges as N/A.

    Source: Table 2 of "VRT: A Video Restoration Transformer" (arXiv:2201.12288)
    Dataset: DAVIS testset
    """

    def __init__(self):
        self.name = "VRT (published)"
        # Published values from Table 2 of VRT paper (DAVIS testset)
        # These are the sigma-specific models
        self.published = {
            10: {'psnr': 38.20, 'ssim': 0.9669},
            20: {'psnr': 35.05, 'ssim': 0.9398},
            30: {'psnr': 33.31, 'ssim': 0.9175},
            40: {'psnr': 32.04, 'ssim': 0.8976},
            50: {'psnr': 31.11, 'ssim': 0.8800},
        }

    def get_results_for_range(self, low, high):
        """
        Check if any published sigma falls within [low, high].
        Returns published values if there's a match, None otherwise.
        """
        for sigma, vals in self.published.items():
            if low <= sigma <= high:
                return vals
        return None


# ============================================================================
#  Benchmark engine
# ============================================================================

class DenoiserBenchmark:
    """
    Benchmark denoisers across noise ranges on DAVIS validation videos.

    The benchmark:
    1. Takes a set of clean video frames from DAVIS
    2. For each noise range, adds Gaussian noise with fixed seed (reproducible)
    3. Runs each denoiser on the same noisy frames
    4. Computes PSNR and SSIM vs clean ground truth
    5. Stores and displays results as a formatted table
    """

    def __init__(self, davis_root, device='cuda', seed=123):
        self.davis_root = davis_root
        self.device = device
        self.seed = seed
        self.denoisers = {}
        self.vrt_published = VRTPublishedResults()

        # Define noise ranges (each is [low, high) for uniform sampling)
        self.noise_ranges = [
            (5, 25), (25, 45), (45, 65), (65, 85), (85, 105),
            (105, 125), (125, 145), (145, 165), (165, 185),
            (185, 205), (205, 225), (225, 245), (245, 255)
        ]

    def register_denoiser(self, name, wrapper):
        """Register a denoiser for benchmarking."""
        self.denoisers[name] = wrapper
        print(f"Registered denoiser: {name}")

    def _load_test_videos(self, max_videos=5, max_frames_per_video=10):
        """
        Load a subset of DAVIS videos for benchmarking.
        Returns dict: {video_name: [list of (H,W,3) float32 [0,1] arrays]}
        """
        videos = {}
        video_dirs = sorted([
            d for d in os.listdir(self.davis_root)
            if os.path.isdir(os.path.join(self.davis_root, d))
        ])

        # Use last N videos (these are likely validation-like if you sorted)
        test_videos = video_dirs[-max_videos:]

        for video_name in test_videos:
            video_path = os.path.join(self.davis_root, video_name)
            frame_files = sorted([
                f for f in os.listdir(video_path)
                if f.endswith(('.png', '.jpg', '.jpeg'))
            ])[:max_frames_per_video]

            frames = []
            for fname in frame_files:
                img = Image.open(os.path.join(video_path, fname)).convert('RGB')
                frames.append(np.array(img, dtype=np.float32) / 255.0)

            if frames:
                videos[video_name] = frames

        total_frames = sum(len(v) for v in videos.values())
        print(f"Loaded {len(videos)} test videos, {total_frames} total frames")
        return videos

    def _add_noise(self, frames, noise_std, rng):
        """Add Gaussian noise to frames. Returns noisy frames and actual std used."""
        noisy = []
        for frame in frames:
            noise = rng.normal(0, noise_std / 255.0, frame.shape).astype(np.float32)
            noisy_frame = np.clip(frame + noise, 0, 1)
            noisy.append(noisy_frame)
        return noisy

    def _compute_metrics(self, clean_frames, processed_frames):
        """Compute average PSNR and SSIM across frames."""
        psnr_vals = []
        ssim_vals = []

        for clean, processed in zip(clean_frames, processed_frames):
            # Convert to uint8 for standard metric computation
            clean_255 = (clean * 255).astype(np.uint8)
            proc_255 = (processed * 255).astype(np.uint8)

            psnr_vals.append(psnr(clean_255, proc_255, data_range=255))
            ssim_vals.append(ssim(clean_255, proc_255, data_range=255, channel_axis=2))

        return {
            'psnr': float(np.mean(psnr_vals)),
            'ssim': float(np.mean(ssim_vals)),
            'psnr_std': float(np.std(psnr_vals)),
            'ssim_std': float(np.std(ssim_vals)),
        }

    def run(self, max_videos=5, max_frames_per_video=10, resize_to=(256, 256),
            results_path=None):
        """
        Run the full benchmark.

        Args:
            max_videos: Number of DAVIS videos to test on
            max_frames_per_video: Max frames per video
            resize_to: Processing resolution for denoisers
            results_path: Path to save JSON results (also loads previous results)

        Returns:
            results: Dict with all benchmark data
        """
        # Load previous results if they exist
        prev_results = {}
        if results_path and os.path.exists(results_path):
            with open(results_path, 'r') as f:
                prev_results = json.load(f)
            print(f"Loaded previous results from {results_path}")

        # Load test videos
        videos = self._load_test_videos(max_videos, max_frames_per_video)

        results = {
            'metadata': {
                'date': datetime.now().isoformat(),
                'davis_root': self.davis_root,
                'num_videos': len(videos),
                'max_frames_per_video': max_frames_per_video,
                'resize_to': list(resize_to),
                'seed': self.seed,
                'noise_ranges': self.noise_ranges,
            },
            'noise_ranges': {},
        }

        print(f"\n{'='*80}")
        print(f"BENCHMARK: Comparing {len(self.denoisers)} denoisers across {len(self.noise_ranges)} noise ranges")
        print(f"{'='*80}\n")

        for low, high in self.noise_ranges:
            range_key = f"{low}-{high}"
            midpoint_sigma = (low + high) / 2.0
            print(f"\n--- Noise range σ = [{low}, {high}] (midpoint: {midpoint_sigma:.0f}) ---")

            # Use fixed seed per range for reproducibility
            rng = np.random.RandomState(self.seed + low)
            noise_std = rng.uniform(low, high)  # Single noise level for this range

            range_results = {
                'noise_std_used': float(noise_std),
                'noisy': None,
                'denoisers': {},
            }

            # Collect all clean and noisy frames across videos
            all_clean = []
            all_noisy = []
            for video_name, clean_frames in videos.items():
                noisy_frames = self._add_noise(clean_frames, noise_std, rng)
                all_clean.extend(clean_frames)
                all_noisy.extend(noisy_frames)

            # Baseline: noisy PSNR/SSIM
            noisy_metrics = self._compute_metrics(all_clean, all_noisy)
            range_results['noisy'] = noisy_metrics
            print(f"  Noisy baseline:  PSNR={noisy_metrics['psnr']:.2f} dB, SSIM={noisy_metrics['ssim']:.4f}")

            # Run each registered denoiser
            for name, wrapper in self.denoisers.items():
                try:
                    t0 = time.time()

                    # Denoise per-video (to maintain temporal context)
                    all_denoised = []
                    frame_offset = 0
                    for video_name, clean_frames in videos.items():
                        n = len(clean_frames)
                        noisy_chunk = all_noisy[frame_offset:frame_offset + n]
                        denoised_chunk = wrapper.denoise(noisy_chunk, noise_std, resize_to)
                        all_denoised.extend(denoised_chunk)
                        frame_offset += n

                    elapsed = time.time() - t0
                    metrics = self._compute_metrics(all_clean, all_denoised)
                    metrics['time_seconds'] = float(elapsed)

                    psnr_gain = metrics['psnr'] - noisy_metrics['psnr']
                    print(f"  {name:20s}: PSNR={metrics['psnr']:.2f} dB (+{psnr_gain:.2f}), "
                          f"SSIM={metrics['ssim']:.4f}, Time={elapsed:.1f}s")

                    range_results['denoisers'][name] = metrics

                except Exception as e:
                    print(f"  {name:20s}: ERROR — {e}")
                    range_results['denoisers'][name] = {'error': str(e)}

            # Add VRT published results
            vrt_vals = self.vrt_published.get_results_for_range(low, high)
            if vrt_vals:
                range_results['denoisers']['VRT (published)'] = {
                    'psnr': vrt_vals['psnr'],
                    'ssim': vrt_vals['ssim'],
                    'note': 'Published values from VRT paper Table 2, DAVIS testset'
                }
                print(f"  {'VRT (published)':20s}: PSNR={vrt_vals['psnr']:.2f} dB, "
                      f"SSIM={vrt_vals['ssim']:.4f} (from paper)")
            else:
                range_results['denoisers']['VRT (published)'] = {
                    'psnr': None, 'ssim': None,
                    'note': f'No published results for sigma range [{low}, {high}]'
                }
                print(f"  {'VRT (published)':20s}: N/A (no published results for this range)")

            results['noise_ranges'][range_key] = range_results

        # Save results
        if results_path:
            # Merge with previous (keep history)
            if 'history' not in prev_results:
                prev_results['history'] = []
            if 'noise_ranges' in prev_results:
                prev_results['history'].append({
                    'date': prev_results.get('metadata', {}).get('date', 'unknown'),
                    'noise_ranges': prev_results['noise_ranges']
                })
            # Save current as latest
            results['history'] = prev_results.get('history', [])

            with open(results_path, 'w') as f:
                json.dump(results, f, indent=2, default=str)
            print(f"\nResults saved to {results_path}")

        # Print formatted table
        self.print_table(results)

        return results

    def print_table(self, results):
        """Print a nicely formatted comparison table."""
        print(f"\n{'='*120}")
        print(f"{'BENCHMARK RESULTS':^120}")
        print(f"{'='*120}")

        # Gather all denoiser names
        all_names = set()
        for range_data in results['noise_ranges'].values():
            all_names.update(range_data['denoisers'].keys())
        all_names = sorted(all_names)

        # Header
        header = f"{'Noise Range':>12} | {'Noisy PSNR':>10} | {'Noisy SSIM':>10}"
        for name in all_names:
            short = name[:12]
            header += f" | {short + ' PSNR':>14} | {short + ' SSIM':>14}"
        print(header)
        print("-" * len(header))

        # Rows
        for range_key in sorted(results['noise_ranges'].keys(),
                                key=lambda x: int(x.split('-')[0])):
            data = results['noise_ranges'][range_key]
            noisy = data['noisy']

            row = f"  σ {range_key:>8} | {noisy['psnr']:>10.2f} | {noisy['ssim']:>10.4f}"

            for name in all_names:
                d = data['denoisers'].get(name, {})
                p = d.get('psnr')
                s = d.get('ssim')
                if p is not None:
                    row += f" | {p:>14.2f} | {s:>14.4f}"
                else:
                    row += f" | {'N/A':>14} | {'N/A':>14}"

            print(row)

        print(f"{'='*120}")
        print(f"Date: {results['metadata']['date']}")
        print(f"Test data: {results['metadata']['num_videos']} videos, "
              f"seed={results['metadata']['seed']}")
        print()

    def save_excel_report(self, results, save_path):
        """
        Save benchmark results as a professionally formatted Excel workbook.
        Contains separate sheets for PSNR, SSIM, PSNR improvement, and a summary.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # Gather denoiser names
        all_names = set()
        for range_data in results['noise_ranges'].values():
            all_names.update(range_data['denoisers'].keys())
        all_names = sorted(all_names)

        sorted_ranges = sorted(
            results['noise_ranges'].keys(),
            key=lambda x: int(x.split('-')[0])
        )

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
        header_fill = PatternFill("solid", fgColor="2F5496")
        subheader_fill = PatternFill("solid", fgColor="D6E4F0")
        subheader_font = Font(bold=True, size=10, name="Arial")
        data_font = Font(size=10, name="Arial")
        na_font = Font(size=10, name="Arial", color="999999", italic=True)
        best_fill = PatternFill("solid", fgColor="C6EFCE")
        best_font = Font(bold=True, size=10, name="Arial", color="006100")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9'),
        )
        center = Alignment(horizontal='center', vertical='center')

        def style_header(ws, row, max_col):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = thin_border

        def style_data_cell(ws, row, col, is_best=False, is_na=False):
            cell = ws.cell(row=row, column=col)
            cell.alignment = center
            cell.border = thin_border
            if is_na:
                cell.font = na_font
            elif is_best:
                cell.font = best_font
                cell.fill = best_fill
            else:
                cell.font = data_font

        # ---- Sheet 1: PSNR Comparison ----
        ws_psnr = wb.active
        ws_psnr.title = "PSNR Comparison"

        # Title
        ws_psnr.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(all_names))
        ws_psnr.cell(1, 1, "PSNR Comparison (dB) — Higher is Better")
        ws_psnr.cell(1, 1).font = Font(bold=True, size=14, name="Arial", color="2F5496")

        # Headers
        headers = ["Noise Range (σ)", "Noisy Baseline"] + all_names
        for c, h in enumerate(headers, 1):
            ws_psnr.cell(3, c, h)
        style_header(ws_psnr, 3, len(headers))

        for r, range_key in enumerate(sorted_ranges, 4):
            data = results['noise_ranges'][range_key]
            ws_psnr.cell(r, 1, range_key)
            ws_psnr.cell(r, 1).font = subheader_font
            ws_psnr.cell(r, 1).alignment = center
            ws_psnr.cell(r, 1).border = thin_border

            ws_psnr.cell(r, 2, round(data['noisy']['psnr'], 2))
            style_data_cell(ws_psnr, r, 2)

            # Find best PSNR for this range
            psnr_vals = {}
            for name in all_names:
                d = data['denoisers'].get(name, {})
                p = d.get('psnr')
                if p is not None:
                    psnr_vals[name] = p
            best_name = max(psnr_vals, key=psnr_vals.get) if psnr_vals else None

            for c, name in enumerate(all_names, 3):
                d = data['denoisers'].get(name, {})
                p = d.get('psnr')
                if p is not None:
                    ws_psnr.cell(r, c, round(p, 2))
                    is_best = (name == best_name)
                    style_data_cell(ws_psnr, r, c, is_best=is_best)
                else:
                    ws_psnr.cell(r, c, "N/A")
                    style_data_cell(ws_psnr, r, c, is_na=True)

        for c in range(1, len(headers) + 1):
            ws_psnr.column_dimensions[get_column_letter(c)].width = 18

        # ---- Sheet 2: SSIM Comparison ----
        ws_ssim = wb.create_sheet("SSIM Comparison")

        ws_ssim.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(all_names))
        ws_ssim.cell(1, 1, "SSIM Comparison — Higher is Better (max 1.0)")
        ws_ssim.cell(1, 1).font = Font(bold=True, size=14, name="Arial", color="2F5496")

        headers_ssim = ["Noise Range (σ)", "Noisy Baseline"] + all_names
        for c, h in enumerate(headers_ssim, 1):
            ws_ssim.cell(3, c, h)
        style_header(ws_ssim, 3, len(headers_ssim))

        for r, range_key in enumerate(sorted_ranges, 4):
            data = results['noise_ranges'][range_key]
            ws_ssim.cell(r, 1, range_key)
            ws_ssim.cell(r, 1).font = subheader_font
            ws_ssim.cell(r, 1).alignment = center
            ws_ssim.cell(r, 1).border = thin_border

            ws_ssim.cell(r, 2, round(data['noisy']['ssim'], 4))
            style_data_cell(ws_ssim, r, 2)

            ssim_vals = {}
            for name in all_names:
                d = data['denoisers'].get(name, {})
                s = d.get('ssim')
                if s is not None:
                    ssim_vals[name] = s
            best_name = max(ssim_vals, key=ssim_vals.get) if ssim_vals else None

            for c, name in enumerate(all_names, 3):
                d = data['denoisers'].get(name, {})
                s = d.get('ssim')
                if s is not None:
                    ws_ssim.cell(r, c, round(s, 4))
                    is_best = (name == best_name)
                    style_data_cell(ws_ssim, r, c, is_best=is_best)
                else:
                    ws_ssim.cell(r, c, "N/A")
                    style_data_cell(ws_ssim, r, c, is_na=True)

        for c in range(1, len(headers_ssim) + 1):
            ws_ssim.column_dimensions[get_column_letter(c)].width = 18

        # ---- Sheet 3: PSNR Improvement over Noisy ----
        ws_gain = wb.create_sheet("PSNR Gain")

        ws_gain.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(all_names))
        ws_gain.cell(1, 1, "PSNR Improvement over Noisy Baseline (dB)")
        ws_gain.cell(1, 1).font = Font(bold=True, size=14, name="Arial", color="2F5496")

        headers_gain = ["Noise Range (σ)"] + all_names
        for c, h in enumerate(headers_gain, 1):
            ws_gain.cell(3, c, h)
        style_header(ws_gain, 3, len(headers_gain))

        for r, range_key in enumerate(sorted_ranges, 4):
            data = results['noise_ranges'][range_key]
            noisy_psnr = data['noisy']['psnr']

            ws_gain.cell(r, 1, range_key)
            ws_gain.cell(r, 1).font = subheader_font
            ws_gain.cell(r, 1).alignment = center
            ws_gain.cell(r, 1).border = thin_border

            gain_vals = {}
            for name in all_names:
                d = data['denoisers'].get(name, {})
                p = d.get('psnr')
                if p is not None:
                    gain_vals[name] = p - noisy_psnr
            best_name = max(gain_vals, key=gain_vals.get) if gain_vals else None

            for c, name in enumerate(all_names, 2):
                d = data['denoisers'].get(name, {})
                p = d.get('psnr')
                if p is not None:
                    gain = p - noisy_psnr
                    ws_gain.cell(r, c, round(gain, 2))
                    is_best = (name == best_name)
                    style_data_cell(ws_gain, r, c, is_best=is_best)
                else:
                    ws_gain.cell(r, c, "N/A")
                    style_data_cell(ws_gain, r, c, is_na=True)

        for c in range(1, len(headers_gain) + 1):
            ws_gain.column_dimensions[get_column_letter(c)].width = 18

        # ---- Sheet 4: Summary & Notes ----
        ws_notes = wb.create_sheet("Notes")
        notes = [
            ["Denoiser Benchmark — Methodology Notes"],
            [""],
            ["Date:", results['metadata']['date']],
            ["Test videos:", str(results['metadata']['num_videos'])],
            ["Frames per video:", str(results['metadata']['max_frames_per_video'])],
            ["Processing resolution:", str(results['metadata']['resize_to'])],
            ["Random seed:", str(results['metadata']['seed'])],
            [""],
            ["Metrics:"],
            ["PSNR", "Peak Signal-to-Noise Ratio. Measures pixel accuracy in dB. Higher = better."],
            ["SSIM", "Structural Similarity Index. Measures perceptual quality (0 to 1). Higher = better."],
            ["PSNR Gain", "Improvement in PSNR over the noisy input (no denoising). Shows how much each denoiser helps."],
            [""],
            ["Model Notes:"],
            ["YourUNet", "Your blind denoiser trained on σ ∈ [5, 255]. Handles all noise levels."],
            ["FastDVDNet", "Trained on σ ∈ [5, 55]. Uses noise sigma as input. Degrades above σ ≈ 55."],
            ["VRT (published)", "Published results from the VRT paper (Table 2). Only available for σ = 10, 20, 30, 40, 50."],
            [""],
            ["Green-highlighted cells indicate the best performer for each noise range."],
        ]
        for r, row_data in enumerate(notes, 1):
            for c, val in enumerate(row_data, 1):
                ws_notes.cell(r, c, val)
                if r == 1:
                    ws_notes.cell(r, c).font = Font(bold=True, size=14, name="Arial", color="2F5496")
                elif c == 1 and r > 2:
                    ws_notes.cell(r, c).font = Font(bold=True, size=10, name="Arial")
                else:
                    ws_notes.cell(r, c).font = data_font

        ws_notes.column_dimensions['A'].width = 22
        ws_notes.column_dimensions['B'].width = 80

        wb.save(save_path)
        print(f"Excel report saved to {save_path}")

    def save_html_report(self, results, save_path):
        """Save benchmark results as a clean HTML report for visual inspection."""
        all_names = set()
        for range_data in results['noise_ranges'].values():
            all_names.update(range_data['denoisers'].keys())
        all_names = sorted(all_names)

        sorted_ranges = sorted(
            results['noise_ranges'].keys(),
            key=lambda x: int(x.split('-')[0])
        )

        def make_table(metric, title, fmt, higher_better=True):
            html = f'<h2>{title}</h2>\n<table>\n<tr><th>Noise Range (σ)</th><th>Noisy Baseline</th>'
            for n in all_names:
                html += f'<th>{n}</th>'
            html += '</tr>\n'

            for rk in sorted_ranges:
                data = results['noise_ranges'][rk]
                noisy_val = data['noisy'].get(metric)
                html += f'<tr><td class="range">{rk}</td><td>{noisy_val:{fmt}}</td>'

                # Find best
                vals = {}
                for name in all_names:
                    d = data['denoisers'].get(name, {})
                    v = d.get(metric)
                    if v is not None:
                        vals[name] = v
                best = max(vals, key=vals.get) if vals else None

                for name in all_names:
                    d = data['denoisers'].get(name, {})
                    v = d.get(metric)
                    if v is not None:
                        cls = ' class="best"' if name == best else ''
                        html += f'<td{cls}>{v:{fmt}}</td>'
                    else:
                        html += '<td class="na">N/A</td>'
                html += '</tr>\n'

            html += '</table>\n'
            return html

        html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>Denoiser Benchmark Results</title>
<style>
    body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 40px; background: #f8f9fa; color: #333; }}
    h1 {{ color: #2F5496; border-bottom: 3px solid #2F5496; padding-bottom: 10px; }}
    h2 {{ color: #2F5496; margin-top: 30px; }}
    table {{ border-collapse: collapse; width: 100%; margin: 15px 0; background: white; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
    th {{ background: #2F5496; color: white; padding: 10px 14px; text-align: center; font-size: 13px; }}
    td {{ padding: 8px 14px; text-align: center; border-bottom: 1px solid #e9ecef; font-size: 13px; }}
    tr:hover {{ background: #f1f3f5; }}
    .range {{ font-weight: bold; background: #e8edf3; }}
    .best {{ background: #c6efce; font-weight: bold; color: #006100; }}
    .na {{ color: #aaa; font-style: italic; }}
    .notes {{ background: white; padding: 20px; border-radius: 6px; margin-top: 30px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
    .notes p {{ margin: 6px 0; line-height: 1.6; }}
    .notes strong {{ color: #2F5496; }}
    .meta {{ color: #666; font-size: 12px; margin-top: 20px; }}
</style></head><body>
<h1>Video Denoiser Benchmark Results</h1>
<p class="meta">Generated: {results['metadata']['date']} | 
{results['metadata']['num_videos']} test videos | 
Seed: {results['metadata']['seed']} | 
Resolution: {results['metadata']['resize_to']}</p>
"""
        html += make_table('psnr', 'PSNR Comparison (dB) — Higher is Better', '.2f')
        html += make_table('ssim', 'SSIM Comparison — Higher is Better', '.4f')

        # PSNR gain table (custom since it's derived)
        html += '<h2>PSNR Improvement over Noisy Baseline (dB)</h2>\n<table>\n'
        html += '<tr><th>Noise Range (σ)</th>'
        for n in all_names:
            html += f'<th>{n}</th>'
        html += '</tr>\n'

        for rk in sorted_ranges:
            data = results['noise_ranges'][rk]
            noisy_p = data['noisy']['psnr']
            html += f'<tr><td class="range">{rk}</td>'

            gains = {}
            for name in all_names:
                d = data['denoisers'].get(name, {})
                p = d.get('psnr')
                if p is not None:
                    gains[name] = p - noisy_p
            best = max(gains, key=gains.get) if gains else None

            for name in all_names:
                d = data['denoisers'].get(name, {})
                p = d.get('psnr')
                if p is not None:
                    g = p - noisy_p
                    cls = ' class="best"' if name == best else ''
                    html += f'<td{cls}>+{g:.2f}</td>'
                else:
                    html += '<td class="na">N/A</td>'
            html += '</tr>\n'
        html += '</table>\n'

        html += """
<div class="notes">
<h2>Notes</h2>
<p><strong>YourUNet:</strong> Blind denoiser trained on σ ∈ [5, 255]. Handles all noise levels without knowing the noise level.</p>
<p><strong>FastDVDNet:</strong> Trained on σ ∈ [5, 55] and requires noise sigma as input. Expected to degrade significantly above σ ≈ 55.</p>
<p><strong>VRT (published):</strong> State-of-the-art transformer model. Results taken from published paper (Table 2, DAVIS testset). Only tested at σ = 10, 20, 30, 40, 50.</p>
<p><strong>Green cells</strong> indicate the best performer for each noise range.</p>
<p><strong>PSNR:</strong> Peak Signal-to-Noise Ratio — measures pixel-level accuracy in decibels. Higher is better.</p>
<p><strong>SSIM:</strong> Structural Similarity Index — measures perceptual quality (0 to 1). Higher is better.</p>
</div>
</body></html>"""

        with open(save_path, 'w') as f:
            f.write(html)
        print(f"HTML report saved to {save_path}")


# ============================================================================
#  Convenience function to run from notebook
# ============================================================================

def run_benchmark(
    your_model,
    davis_root,
    fastdvdnet_weights_path=None,
    device='cuda',
    max_videos=5,
    max_frames_per_video=10,
    resize_to=(256, 256),
    save_dir='./denoiser_evaluation'
):
    """
    Run the full benchmark from your training notebook.

    Args:
        your_model: Your trained BlindVideoDenoiserUNet (already loaded, eval mode)
        davis_root: Path to DAVIS dataset folder
        fastdvdnet_weights_path: Path to FastDVDNet model.pth (will skip if None)
        device: 'cuda' or 'cpu'
        max_videos: Number of test videos
        max_frames_per_video: Frames per video
        resize_to: Processing resolution
        save_dir: Directory to save all outputs (JSON, Excel, HTML)

    Returns:
        results dict

    Usage in notebook:
        from benchmark import run_benchmark
        results = run_benchmark(
            your_model=model,
            davis_root='/content/DAVISDataset',
            fastdvdnet_weights_path='/content/fastdvdnet/model.pth',
            save_dir='/content/drive/MyDrive/ResearchProject/denoiser_evaluation'
        )
    """
    save_dir = Path(save_dir)
    save_dir.mkdir(parents=True, exist_ok=True)

    results_json = save_dir / "results_data.json"
    results_xlsx = save_dir / "comparison_tables.xlsx"
    results_html = save_dir / "comparison_report.html"

    benchmark = DenoiserBenchmark(davis_root, device=device)

    # Register your model
    your_wrapper = YourDenoiserWrapper(your_model, device=device)
    benchmark.register_denoiser("YourUNet", your_wrapper)

    # Register FastDVDNet if weights available
    if fastdvdnet_weights_path and os.path.exists(fastdvdnet_weights_path):
        try:
            fdvd_wrapper = FastDVDNetWrapper(fastdvdnet_weights_path, device=device)
            benchmark.register_denoiser("FastDVDNet", fdvd_wrapper)
        except Exception as e:
            print(f"Warning: Could not load FastDVDNet: {e}")
            print("Continuing without FastDVDNet...")
    else:
        print(f"FastDVDNet weights not found at {fastdvdnet_weights_path}")
        print("To include FastDVDNet, download from: https://github.com/m-tassano/fastdvdnet")
        print("  !git clone https://github.com/m-tassano/fastdvdnet /content/fastdvdnet")

    # Run benchmark
    results = benchmark.run(
        max_videos=max_videos,
        max_frames_per_video=max_frames_per_video,
        resize_to=resize_to,
        results_path=str(results_json),
    )

    # Save Excel and HTML reports
    benchmark.save_excel_report(results, str(results_xlsx))
    benchmark.save_html_report(results, str(results_html))

    print(f"\n{'='*60}")
    print(f"All outputs saved to: {save_dir}")
    print(f"  - {results_json.name:30s} (raw data, accumulates history)")
    print(f"  - {results_xlsx.name:30s} (formatted spreadsheet)")
    print(f"  - {results_html.name:30s} (visual report)")
    print(f"{'='*60}")

    return results


if __name__ == "__main__":
    print("=== Denoiser Benchmark ===")
    print()
    print("Usage in your Colab notebook:")
    print()
    print("  # 1. Download FastDVDNet pretrained weights")
    print("  !git clone https://github.com/m-tassano/fastdvdnet /content/fastdvdnet")
    print()
    print("  # 2. Load your trained model")
    print("  from unet_denoiser import BlindVideoDenoiserUNet")
    print("  model = BlindVideoDenoiserUNet(in_channels=9, out_channels=3, base_channels=64, num_stages=3)")
    print("  ckpt = torch.load('/content/checkpoints/best_model.pt', map_location='cuda')")
    print("  model.load_state_dict(ckpt['model_state_dict'])")
    print("  model.eval()")
    print()
    print("  # 3. Run benchmark (saves JSON + Excel + HTML to the folder)")
    print("  from benchmark import run_benchmark")
    print("  results = run_benchmark(")
    print("      your_model=model,")
    print("      davis_root='/content/DAVISDataset',")
    print("      fastdvdnet_weights_path='/content/fastdvdnet/model.pth',")
    print("      save_dir='/content/drive/MyDrive/ResearchProject/denoiser_evaluation'")
    print("  )")
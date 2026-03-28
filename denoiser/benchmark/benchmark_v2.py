"""
Comprehensive Benchmark for Blind Video Denoisers
==================================================
Compares your UNet denoiser against FastDVDNet and VRT across noise ranges.

Metrics:
- PSNR (Peak Signal-to-Noise Ratio): Measures pixel-level accuracy in dB.
  Computed as 10 * log10(MAX^2 / MSE). Higher = better. ~30dB is good, ~40dB is excellent.
  Directly reflects how well the MSE loss was minimized during training.

- SSIM (Structural Similarity Index): Measures perceptual quality (0 to 1).
  Compares luminance, contrast, and local structure between patches.
  Unlike PSNR, SSIM penalizes blurriness and structural distortion. 1.0 = identical.

How benchmarking works:
1. For each noise range (e.g., 5-25, 25-45, ...):
   - Take N frames from DAVIS validation videos
   - Add Gaussian noise with std uniformly sampled from that range
   - Run each denoiser on the same noisy input
   - Compute PSNR and SSIM of denoised output vs. clean ground truth
   - The noisy PSNR/SSIM is also recorded as a baseline (no denoising)

2. All denoisers see the exact same noisy frames (seeded RNG), so comparison is fair.

3. Results are stored as a JSON and printed as a formatted table.

Important notes on the comparison models:
- FastDVDNet was trained on noise sigma [5, 55]. It will degrade significantly
  above sigma 55 since it has never seen such noise levels during training.
- VRT was trained on fixed sigma values (e.g., sigma=30, 50). Similar limitation.
- This denoiser model was trained on sigma [5, 255], so it should handle all ranges.
  The benchmark will clearly show where each model's training range matters.
"""

import torch
import torch.nn as nn
import numpy as np
import os
import json
import time
from pathlib import Path
from PIL import Image
from skimage.metrics import peak_signal_noise_ratio as psnr
from skimage.metrics import structural_similarity as ssim
from datetime import datetime


# ============================================================================
#  Denoiser wrappers (unified interface for benchmarking)
# ============================================================================

class YourDenoiserWrapper:
    """Wrapper for your BlindVideoDenoiserUNet (supports 3 or 5 frame temporal input)."""

    def __init__(self, model, device='cuda', num_input_frames=5):
        self.model = model.to(device).eval()
        self.device = device
        self.name = "YourUNet"
        self.num_input_frames = num_input_frames
        self.half_window = num_input_frames // 2

    @torch.no_grad()
    def denoise(self, noisy_frames, noise_sigma, resize_to=(256, 256)):
        """
        Args:
            noisy_frames: list of numpy arrays (H, W, 3) float32 [0, 1]
                          Already at the target resolution (pre-resized by benchmark).
            noise_sigma: noise std (0-255 scale, not used by model — it's blind)
            resize_to: target resolution (used only if frames differ from this)
        Returns:
            list of denoised numpy arrays (H, W, 3) float32 [0, 1]
        """
        n = len(noisy_frames)
        denoised = []

        for i in range(n):
            # Gather temporal neighbors with boundary handling
            neighbor_frames = []
            for offset in range(-self.half_window, self.half_window + 1):
                idx = max(0, min(n - 1, i + offset))
                frame = noisy_frames[idx]
                if frame.shape[0] != resize_to[0] or frame.shape[1] != resize_to[1]:
                    frame = self._resize(frame, resize_to)
                neighbor_frames.append(frame)

            # Concatenate: (H, W, num_frames*3)
            concat = np.concatenate(neighbor_frames, axis=2)
            input_t = torch.from_numpy(concat).permute(2, 0, 1).unsqueeze(0).float().to(self.device)

            out = self.model(input_t)
            out = torch.clamp(out.squeeze(0), 0, 1).cpu().permute(1, 2, 0).numpy()
            denoised.append(out)

        return denoised

    def _resize(self, img_np, size):
        """Resize (H, W, 3) float32 [0,1] numpy array."""
        pil = Image.fromarray((img_np * 255).astype(np.uint8))
        pil = pil.resize((size[1], size[0]), Image.BILINEAR)
        return np.array(pil, dtype=np.float32) / 255.0


# ============================================================================
#  FastDVDNet — uses official repo code (cloned to /content/fastdvdnet)
# ============================================================================

class FastDVDNetWrapper:
    """
    Wrapper for FastDVDNet using the official repo's model definition.
    Requires: !git clone https://github.com/m-tassano/fastdvdnet /content/fastdvdnet
    """

    def __init__(self, repo_path, model_filename='model.pth', device='cuda'):
        import sys
        self.device = device
        self.name = "FastDVDNet"

        # Add the repo to sys.path so we can import their model
        if repo_path not in sys.path:
            sys.path.insert(0, repo_path)

        from models import FastDVDnet as OfficialFastDVDnet

        self.model = OfficialFastDVDnet(num_input_frames=5)

        # Load pretrained weights
        model_path = os.path.join(repo_path, model_filename)
        state_dict = torch.load(model_path, map_location=device)

        # Handle DataParallel 'module.' prefix
        if any(k.startswith('module.') for k in state_dict.keys()):
            state_dict = {k.replace('module.', ''): v for k, v in state_dict.items()}

        self.model.load_state_dict(state_dict)
        self.model = self.model.to(device).eval()
        print(f"FastDVDNet loaded from {model_path}")

        # Also import their denoising function
        try:
            from fastdvdnet import denoise_seq_fastdvdnet
            self.denoise_seq_fn = denoise_seq_fastdvdnet
            self.use_official_fn = True
        except ImportError:
            self.use_official_fn = False

    @torch.no_grad()
    def denoise(self, noisy_frames, noise_sigma, resize_to=(256, 256)):
        """
        Args:
            noisy_frames: list of numpy arrays (H, W, 3) float32 [0, 1]
                          Already at the target resolution (pre-resized by benchmark).
            noise_sigma: noise std (0-255 scale) — FastDVDNet uses this as input
            resize_to: target resolution (used only if frames differ)
        Returns:
            list of denoised numpy arrays (H, W, 3) float32 [0, 1]
        """
        # Resize only if needed
        frames_ready = []
        for f in noisy_frames:
            if f.shape[0] != resize_to[0] or f.shape[1] != resize_to[1]:
                frames_ready.append(self._resize(f, resize_to))
            else:
                frames_ready.append(f)

        # Stack to (N, C, H, W) tensor
        seq = torch.stack([
            torch.from_numpy(f).permute(2, 0, 1) for f in frames_ready
        ]).float().to(self.device)

        # Noise sigma as scalar tensor (normalized to 0-1 range)
        sigma_tensor = torch.FloatTensor([noise_sigma / 255.0]).to(self.device)

        if self.use_official_fn:
            # Use their official denoising function which handles boundary frames
            out_seq = self.denoise_seq_fn(
                seq=seq,
                noise_std=sigma_tensor,
                temp_psz=5,
                model_temporal=self.model
            )
        else:
            # Manual frame-by-frame fallback
            n = len(seq)
            h, w = seq.shape[2], seq.shape[3]
            out_seq = torch.empty_like(seq)
            for i in range(n):
                indices = [
                    max(0, i - 2), max(0, i - 1), i,
                    min(n - 1, i + 1), min(n - 1, i + 2)
                ]
                in_frames = seq[indices].unsqueeze(0)  # (1, 5, C, H, W)
                noise_map = sigma_tensor.expand(1, 1, h, w)
                out_seq[i] = self.model(in_frames, noise_map).squeeze(0)

        # Convert back to numpy (no resize needed — already at target resolution)
        denoised = []
        for i in range(len(noisy_frames)):
            frame = torch.clamp(out_seq[i], 0, 1).cpu().permute(1, 2, 0).numpy()
            denoised.append(frame)

        return denoised

    def _resize(self, img_np, size):
        pil = Image.fromarray((img_np * 255).astype(np.uint8))
        pil = pil.resize((size[1], size[0]), Image.BILINEAR)
        return np.array(pil, dtype=np.float32) / 255.0


class VRTPublishedResults:
    """
    Published VRT denoising results from Table VII of the VRT paper (DAVIS testset).
    Stored by sigma, matched to benchmark PSNR levels by finding the closest one.
    """

    def __init__(self):
        self.name = "VRT (published)"
        # Published values from Table VII of VRT paper (DAVIS testset)
        # SSIM not published — marked as None
        # noisy_psnr = 20*log10(255/sigma), the theoretical noisy baseline
        self.published = {
            10: {'psnr': 40.82, 'ssim': None, 'noisy_psnr': 28.1},
            20: {'psnr': 38.15, 'ssim': None, 'noisy_psnr': 22.1},
            30: {'psnr': 36.52, 'ssim': None, 'noisy_psnr': 18.6},
            40: {'psnr': 35.32, 'ssim': None, 'noisy_psnr': 16.1},
            50: {'psnr': 34.36, 'ssim': None, 'noisy_psnr': 14.1},
        }

    def get_results_for_psnr(self, target_psnr, tolerance=1.0):
        """
        Return published values if a published sigma's noisy PSNR is within
        tolerance of the target PSNR level. Returns (result_dict, sigma) or (None, None).
        """
        best_match = None
        best_sigma = None
        best_diff = float('inf')
        for sigma, vals in self.published.items():
            diff = abs(vals['noisy_psnr'] - target_psnr)
            if diff < best_diff and diff <= tolerance:
                best_diff = diff
                best_match = vals
                best_sigma = sigma
        return best_match, best_sigma


class FastDVDNetPublishedResults:
    """
    Published FastDVDNet results from Table VII of VRT paper (DAVIS testset).
    Used as a sanity check for live FastDVDNet results.
    """

    def __init__(self):
        self.name = "FastDVDNet (published)"
        self.published = {
            10: {'psnr': 38.71, 'ssim': None, 'noisy_psnr': 28.1},
            20: {'psnr': 35.77, 'ssim': None, 'noisy_psnr': 22.1},
            30: {'psnr': 34.04, 'ssim': None, 'noisy_psnr': 18.6},
            40: {'psnr': 32.82, 'ssim': None, 'noisy_psnr': 16.1},
            50: {'psnr': 31.86, 'ssim': None, 'noisy_psnr': 14.1},
        }

    def get_results_for_psnr(self, target_psnr, tolerance=1.0):
        best_match = None
        best_sigma = None
        best_diff = float('inf')
        for sigma, vals in self.published.items():
            diff = abs(vals['noisy_psnr'] - target_psnr)
            if diff < best_diff and diff <= tolerance:
                best_diff = diff
                best_match = vals
                best_sigma = sigma
        return best_match, best_sigma

class DenoiserBenchmark:
    """
    Benchmark denoisers across noise ranges on DAVIS validation videos.

    The benchmark:
    1. Takes a set of clean video frames from DAVIS
    2. For each noise range, adds Gaussian noise with fixed seed (reproducible)
    3. Runs each denoiser on the same noisy frames
    4. Computes PSNR and SSIM vs clean ground truth
    5. Stores and displays results as a formatted table
    """

    def __init__(self, davis_root, device='cuda', seed=123):
        self.davis_root = davis_root
        self.device = device
        self.seed = seed
        self.denoisers = {}
        self.vrt_published = VRTPublishedResults()
        self.fastdvdnet_published = FastDVDNetPublishedResults()

        # Noise levels to test:
        # - PSNR-based levels for uniform perceptual coverage
        # - VRT/FastDVDNet exact sigmas are included via published results
        #   matched to the nearest PSNR level
        #
        # Format: list of target PSNR values (dB).
        # σ is derived as: σ = 255 / 10^(PSNR/20)
        self.target_psnrs = [5, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30, 34, 38]

        # VRT published results are at specific sigmas; map them to PSNR
        # σ=10 → 28.1 dB, σ=20 → 22.1 dB, σ=30 → 18.6 dB, σ=40 → 16.1 dB, σ=50 → 14.1 dB
        # These are close to our PSNR grid points (28, 22, 18, 16, 14)

    def register_denoiser(self, name, wrapper):
        """Register a denoiser for benchmarking."""
        self.denoisers[name] = wrapper
        print(f"Registered denoiser: {name}")

    def _load_test_videos(self, max_videos=5, max_frames_per_video=10, resize_to=None):
        """
        Load a subset of DAVIS videos for benchmarking.
        If resize_to is provided, frames are resized at load time so that
        noise addition, denoising, and metric computation all happen at the
        same resolution — avoiding interpolation artifacts in the metrics.

        Returns dict: {video_name: [list of (H,W,3) float32 [0,1] arrays]}
        """
        videos = {}
        video_dirs = sorted([
            d for d in os.listdir(self.davis_root)
            if os.path.isdir(os.path.join(self.davis_root, d))
        ])

        # Use last N videos (these are likely validation-like if you sorted)
        test_videos = video_dirs[-max_videos:]

        for video_name in test_videos:
            video_path = os.path.join(self.davis_root, video_name)
            frame_files = sorted([
                f for f in os.listdir(video_path)
                if f.endswith(('.png', '.jpg', '.jpeg'))
            ])[:max_frames_per_video]

            frames = []
            for fname in frame_files:
                img = Image.open(os.path.join(video_path, fname)).convert('RGB')
                if resize_to is not None:
                    img = img.resize((resize_to[1], resize_to[0]), Image.LANCZOS)
                frames.append(np.array(img, dtype=np.float32) / 255.0)

            if frames:
                videos[video_name] = frames

        total_frames = sum(len(v) for v in videos.values())
        res_str = f"{resize_to[1]}x{resize_to[0]}" if resize_to else "native"
        print(f"Loaded {len(videos)} test videos, {total_frames} total frames ({res_str})")
        return videos

    def _add_noise(self, frames, noise_std, rng):
        """Add Gaussian noise to frames. Returns noisy frames and actual std used."""
        noisy = []
        for frame in frames:
            noise = rng.normal(0, noise_std / 255.0, frame.shape).astype(np.float32)
            noisy_frame = np.clip(frame + noise, 0, 1)
            noisy.append(noisy_frame)
        return noisy

    def _compute_metrics(self, clean_frames, processed_frames):
        """Compute average PSNR and SSIM across frames."""
        psnr_vals = []
        ssim_vals = []

        for clean, processed in zip(clean_frames, processed_frames):
            # Convert to uint8 for standard metric computation
            clean_255 = (clean * 255).astype(np.uint8)
            proc_255 = (processed * 255).astype(np.uint8)

            psnr_vals.append(psnr(clean_255, proc_255, data_range=255))
            ssim_vals.append(ssim(clean_255, proc_255, data_range=255, channel_axis=2))

        return {
            'psnr': float(np.mean(psnr_vals)),
            'ssim': float(np.mean(ssim_vals)),
            'psnr_std': float(np.std(psnr_vals)),
            'ssim_std': float(np.std(ssim_vals)),
        }

    def run(self, max_videos=5, max_frames_per_video=10, resize_to=(480, 856),
            results_path=None):
        """
        Run the full benchmark.

        Args:
            max_videos: Number of DAVIS videos to test on
            max_frames_per_video: Max frames per video
            resize_to: Processing resolution for denoisers
            results_path: Path to save JSON results (also loads previous results)

        Returns:
            results: Dict with all benchmark data
        """
        # Load previous results if they exist
        prev_results = {}
        if results_path and os.path.exists(results_path):
            with open(results_path, 'r') as f:
                prev_results = json.load(f)
            print(f"Loaded previous results from {results_path}")

        # Load test videos — resized to benchmark resolution so that
        # noise, denoising, and metrics all happen at the same resolution
        videos = self._load_test_videos(max_videos, max_frames_per_video, resize_to=resize_to)

        results = {
            'metadata': {
                'date': datetime.now().isoformat(),
                'davis_root': self.davis_root,
                'num_videos': len(videos),
                'max_frames_per_video': max_frames_per_video,
                'resize_to': list(resize_to),
                'seed': self.seed,
                'target_psnrs': self.target_psnrs,
            },
            'noise_ranges': {},
        }

        print(f"\n{'='*80}")
        print(f"BENCHMARK: Comparing {len(self.denoisers)} denoisers across {len(self.target_psnrs)} noise levels")
        print(f"{'='*80}\n")

        for target_psnr in self.target_psnrs:
            # Convert PSNR to sigma: σ = 255 / 10^(PSNR/20)
            noise_std = 255.0 / (10 ** (target_psnr / 20.0))
            range_key = f"PSNR={target_psnr}dB"

            print(f"\n--- Noisy PSNR ≈ {target_psnr} dB (σ ≈ {noise_std:.1f}) ---")

            # Use fixed seed per level for reproducibility
            rng = np.random.RandomState(self.seed + target_psnr)

            range_results = {
                'target_psnr': target_psnr,
                'noise_std_used': float(noise_std),
                'noisy': None,
                'denoisers': {},
            }

            # Collect all clean and noisy frames across videos
            all_clean = []
            all_noisy = []
            for video_name, clean_frames in videos.items():
                noisy_frames = self._add_noise(clean_frames, noise_std, rng)
                all_clean.extend(clean_frames)
                all_noisy.extend(noisy_frames)

            # Baseline: noisy PSNR/SSIM
            noisy_metrics = self._compute_metrics(all_clean, all_noisy)
            range_results['noisy'] = noisy_metrics
            print(f"  Noisy baseline:  PSNR={noisy_metrics['psnr']:.2f} dB, SSIM={noisy_metrics['ssim']:.4f}")

            # Run each registered denoiser
            for name, wrapper in self.denoisers.items():
                try:
                    t0 = time.time()

                    # Denoise per-video (to maintain temporal context)
                    all_denoised = []
                    frame_offset = 0
                    for video_name, clean_frames in videos.items():
                        n = len(clean_frames)
                        noisy_chunk = all_noisy[frame_offset:frame_offset + n]
                        denoised_chunk = wrapper.denoise(noisy_chunk, noise_std, resize_to)
                        all_denoised.extend(denoised_chunk)
                        frame_offset += n

                    elapsed = time.time() - t0
                    metrics = self._compute_metrics(all_clean, all_denoised)
                    metrics['time_seconds'] = float(elapsed)

                    psnr_gain = metrics['psnr'] - noisy_metrics['psnr']
                    print(f"  {name:20s}: PSNR={metrics['psnr']:.2f} dB (+{psnr_gain:.2f}), "
                          f"SSIM={metrics['ssim']:.4f}, Time={elapsed:.1f}s")

                    range_results['denoisers'][name] = metrics

                except Exception as e:
                    print(f"  {name:20s}: ERROR — {e}")
                    range_results['denoisers'][name] = {'error': str(e)}

            # Add VRT published results (matched by PSNR proximity)
            vrt_vals, vrt_sigma = self.vrt_published.get_results_for_psnr(target_psnr)
            if vrt_vals:
                range_results['denoisers']['VRT (published)'] = {
                    'psnr': vrt_vals['psnr'],
                    'ssim': vrt_vals['ssim'],
                    'note': f'Published at σ={vrt_sigma}, Table VII, DAVIS'
                }
                print(f"  {'VRT (published)':20s}: PSNR={vrt_vals['psnr']:.2f} dB (from paper, σ={vrt_sigma})")
            else:
                range_results['denoisers']['VRT (published)'] = {
                    'psnr': None, 'ssim': None,
                    'note': 'No published results near this noise level'
                }
                print(f"  {'VRT (published)':20s}: N/A")

            # Add FastDVDNet published results
            fdvd_vals, fdvd_sigma = self.fastdvdnet_published.get_results_for_psnr(target_psnr)
            if fdvd_vals:
                range_results['denoisers']['FastDVDNet (published)'] = {
                    'psnr': fdvd_vals['psnr'],
                    'ssim': fdvd_vals['ssim'],
                    'note': f'Published at σ={fdvd_sigma}, Table VII, DAVIS'
                }
                print(f"  {'FastDVDNet (pub.)':20s}: PSNR={fdvd_vals['psnr']:.2f} dB (from paper, σ={fdvd_sigma})")
            else:
                range_results['denoisers']['FastDVDNet (published)'] = {
                    'psnr': None, 'ssim': None,
                    'note': 'No published results near this noise level'
                }
                print(f"  {'FastDVDNet (pub.)':20s}: N/A")

            results['noise_ranges'][range_key] = range_results

        # Save results
        if results_path:
            # Merge with previous (keep history)
            if 'history' not in prev_results:
                prev_results['history'] = []
            if 'noise_ranges' in prev_results:
                prev_results['history'].append({
                    'date': prev_results.get('metadata', {}).get('date', 'unknown'),
                    'noise_ranges': prev_results['noise_ranges']
                })
            # Save current as latest
            results['history'] = prev_results.get('history', [])

            with open(results_path, 'w') as f:
                json.dump(results, f, indent=2, default=str)
            print(f"\nResults saved to {results_path}")

        # Print formatted table
        self.print_table(results)

        return results

    def print_table(self, results):
        """Print a nicely formatted comparison table."""
        print(f"\n{'='*120}")
        print(f"{'BENCHMARK RESULTS':^120}")
        print(f"{'='*120}")

        # Gather all denoiser names
        all_names = set()
        for range_data in results['noise_ranges'].values():
            all_names.update(range_data['denoisers'].keys())
        all_names = sorted(all_names)

        # Header
        header = f"{'Noise Level':>16} | {'σ':>6} | {'Noisy PSNR':>10} | {'Noisy SSIM':>10}"
        for name in all_names:
            short = name[:12]
            header += f" | {short + ' PSNR':>14} | {short + ' SSIM':>14}"
        print(header)
        print("-" * len(header))

        def sort_key(x):
            """Sort 'PSNR=10dB' keys numerically (descending PSNR = ascending noise)."""
            if x.startswith('PSNR='):
                return int(x.split('=')[1].replace('dB', ''))
            if x.startswith('σ='):
                return int(x.split('=')[1])
            return int(x.split('-')[0])

        # Rows — sort by PSNR descending (low noise first)
        for range_key in sorted(results['noise_ranges'].keys(), key=sort_key, reverse=True):
            data = results['noise_ranges'][range_key]
            noisy = data['noisy']
            sigma = data.get('noise_std_used', 0)

            row = f"  {range_key:>14} | {sigma:>6.1f} | {noisy['psnr']:>10.2f} | {noisy['ssim']:>10.4f}"

            for name in all_names:
                d = data['denoisers'].get(name, {})
                p = d.get('psnr')
                s = d.get('ssim')
                if p is not None:
                    s_str = f"{s:>14.4f}" if s is not None else f"{'N/A':>14}"
                    row += f" | {p:>14.2f} | {s_str}"
                else:
                    row += f" | {'N/A':>14} | {'N/A':>14}"

            print(row)

        print(f"{'='*120}")
        print(f"Date: {results['metadata']['date']}")
        print(f"Test data: {results['metadata']['num_videos']} videos, "
              f"seed={results['metadata']['seed']}")
        print()

    def save_excel_report(self, results, save_path):
        """
        Save benchmark results as a professionally formatted Excel workbook.
        Contains separate sheets for PSNR, SSIM, PSNR improvement, and a summary.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # Gather denoiser names
        all_names = set()
        for range_data in results['noise_ranges'].values():
            all_names.update(range_data['denoisers'].keys())
        all_names = sorted(all_names)

        def sort_key(x):
            if x.startswith('PSNR='):
                return int(x.split('=')[1].replace('dB', ''))
            if x.startswith('σ='):
                return int(x.split('=')[1])
            return int(x.split('-')[0])

        sorted_ranges = sorted(
            results['noise_ranges'].keys(),
            key=sort_key,
            reverse=True  # High PSNR (low noise) first
        )
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
        header_fill = PatternFill("solid", fgColor="2F5496")
        subheader_fill = PatternFill("solid", fgColor="D6E4F0")
        subheader_font = Font(bold=True, size=10, name="Arial")
        data_font = Font(size=10, name="Arial")
        na_font = Font(size=10, name="Arial", color="999999", italic=True)
        best_fill = PatternFill("solid", fgColor="C6EFCE")
        best_font = Font(bold=True, size=10, name="Arial", color="006100")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9'),
        )
        center = Alignment(horizontal='center', vertical='center')

        def style_header(ws, row, max_col):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = thin_border

        def style_data_cell(ws, row, col, is_best=False, is_na=False):
            cell = ws.cell(row=row, column=col)
            cell.alignment = center
            cell.border = thin_border
            if is_na:
                cell.font = na_font
            elif is_best:
                cell.font = best_font
                cell.fill = best_fill
            else:
                cell.font = data_font

        # ---- Sheet 1: PSNR Comparison ----
        ws_psnr = wb.active
        ws_psnr.title = "PSNR Comparison"

        # Title
        ws_psnr.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(all_names))
        ws_psnr.cell(1, 1, "PSNR Comparison (dB)")
        ws_psnr.cell(1, 1).font = Font(bold=True, size=14, name="Arial", color="2F5496")

        # Headers
        headers = ["Noise Level", "σ", "Noisy Baseline"] + all_names
        for c, h in enumerate(headers, 1):
            ws_psnr.cell(3, c, h)
        style_header(ws_psnr, 3, len(headers))

        for r, range_key in enumerate(sorted_ranges, 4):
            data = results['noise_ranges'][range_key]
            ws_psnr.cell(r, 1, range_key)
            ws_psnr.cell(r, 1).font = subheader_font
            ws_psnr.cell(r, 1).alignment = center
            ws_psnr.cell(r, 1).border = thin_border

            sigma = data.get('noise_std_used', 0)
            ws_psnr.cell(r, 2, round(sigma, 1))
            style_data_cell(ws_psnr, r, 2)

            ws_psnr.cell(r, 3, round(data['noisy']['psnr'], 2))
            style_data_cell(ws_psnr, r, 3)

            # Find best PSNR for this range
            psnr_vals = {}
            for name in all_names:
                d = data['denoisers'].get(name, {})
                p = d.get('psnr')
                if p is not None:
                    psnr_vals[name] = p
            best_name = max(psnr_vals, key=psnr_vals.get) if psnr_vals else None

            for c, name in enumerate(all_names, 4):
                d = data['denoisers'].get(name, {})
                p = d.get('psnr')
                if p is not None:
                    ws_psnr.cell(r, c, round(p, 2))
                    is_best = (name == best_name)
                    style_data_cell(ws_psnr, r, c, is_best=is_best)
                else:
                    ws_psnr.cell(r, c, "N/A")
                    style_data_cell(ws_psnr, r, c, is_na=True)

        for c in range(1, len(headers) + 1):
            ws_psnr.column_dimensions[get_column_letter(c)].width = 18

        # ---- Sheet 2: SSIM Comparison ----
        ws_ssim = wb.create_sheet("SSIM Comparison")

        ws_ssim.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(all_names))
        ws_ssim.cell(1, 1, "SSIM Comparison — Higher is Better (max 1.0)")
        ws_ssim.cell(1, 1).font = Font(bold=True, size=14, name="Arial", color="2F5496")

        headers_ssim = ["Noise Level", "σ", "Noisy Baseline"] + all_names
        for c, h in enumerate(headers_ssim, 1):
            ws_ssim.cell(3, c, h)
        style_header(ws_ssim, 3, len(headers_ssim))

        for r, range_key in enumerate(sorted_ranges, 4):
            data = results['noise_ranges'][range_key]
            ws_ssim.cell(r, 1, range_key)
            ws_ssim.cell(r, 1).font = subheader_font
            ws_ssim.cell(r, 1).alignment = center
            ws_ssim.cell(r, 1).border = thin_border

            sigma = data.get('noise_std_used', 0)
            ws_ssim.cell(r, 2, round(sigma, 1))
            style_data_cell(ws_ssim, r, 2)

            ws_ssim.cell(r, 3, round(data['noisy']['ssim'], 4))
            style_data_cell(ws_ssim, r, 3)

            ssim_vals = {}
            for name in all_names:
                d = data['denoisers'].get(name, {})
                s = d.get('ssim')
                if s is not None:
                    ssim_vals[name] = s
            best_name = max(ssim_vals, key=ssim_vals.get) if ssim_vals else None

            for c, name in enumerate(all_names, 4):
                d = data['denoisers'].get(name, {})
                s = d.get('ssim')
                if s is not None:
                    ws_ssim.cell(r, c, round(s, 4))
                    is_best = (name == best_name)
                    style_data_cell(ws_ssim, r, c, is_best=is_best)
                else:
                    ws_ssim.cell(r, c, "N/A")
                    style_data_cell(ws_ssim, r, c, is_na=True)

        for c in range(1, len(headers_ssim) + 1):
            ws_ssim.column_dimensions[get_column_letter(c)].width = 18

        # ---- Sheet 3: PSNR Improvement over Noisy ----
        ws_gain = wb.create_sheet("PSNR Gain")

        ws_gain.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(all_names))
        ws_gain.cell(1, 1, "PSNR Improvement over Noisy Baseline (dB)")
        ws_gain.cell(1, 1).font = Font(bold=True, size=14, name="Arial", color="2F5496")

        headers_gain = ["Noise Level", "σ"] + all_names
        for c, h in enumerate(headers_gain, 1):
            ws_gain.cell(3, c, h)
        style_header(ws_gain, 3, len(headers_gain))

        for r, range_key in enumerate(sorted_ranges, 4):
            data = results['noise_ranges'][range_key]
            noisy_psnr = data['noisy']['psnr']

            ws_gain.cell(r, 1, range_key)
            ws_gain.cell(r, 1).font = subheader_font
            ws_gain.cell(r, 1).alignment = center
            ws_gain.cell(r, 1).border = thin_border

            sigma = data.get('noise_std_used', 0)
            ws_gain.cell(r, 2, round(sigma, 1))
            style_data_cell(ws_gain, r, 2)

            gain_vals = {}
            for name in all_names:
                d = data['denoisers'].get(name, {})
                p = d.get('psnr')
                if p is not None:
                    gain_vals[name] = p - noisy_psnr
            best_name = max(gain_vals, key=gain_vals.get) if gain_vals else None

            for c, name in enumerate(all_names, 3):
                d = data['denoisers'].get(name, {})
                p = d.get('psnr')
                if p is not None:
                    gain = p - noisy_psnr
                    ws_gain.cell(r, c, round(gain, 2))
                    is_best = (name == best_name)
                    style_data_cell(ws_gain, r, c, is_best=is_best)
                else:
                    ws_gain.cell(r, c, "N/A")
                    style_data_cell(ws_gain, r, c, is_na=True)

        for c in range(1, len(headers_gain) + 1):
            ws_gain.column_dimensions[get_column_letter(c)].width = 18

        # ---- Sheet 4: Summary & Notes ----
        ws_notes = wb.create_sheet("Notes")
        notes = [
            ["Denoiser Benchmark — Methodology Notes"],
            [""],
            ["Date:", results['metadata']['date']],
            ["Test videos:", str(results['metadata']['num_videos'])],
            ["Frames per video:", str(results['metadata']['max_frames_per_video'])],
            ["Processing resolution:", str(results['metadata']['resize_to'])],
            ["Random seed:", str(results['metadata']['seed'])],
            [""],
            ["Metrics:"],
            ["PSNR", "Peak Signal-to-Noise Ratio. Measures pixel accuracy in dB. Higher = better."],
            ["SSIM", "Structural Similarity Index. Measures perceptual quality (0 to 1). Higher = better."],
            ["PSNR Gain", "Improvement in PSNR over the noisy input (no denoising). Shows how much each denoiser helps."],
            [""],
            ["Model Notes:"],
            ["YourUNet", "Your blind denoiser. Noise sampled uniformly in PSNR space for balanced training."],
            ["FastDVDNet", "Trained on σ ∈ [5, 55]. Uses noise sigma as input. Degrades above σ ≈ 55."],
            ["VRT (published)", "Published PSNR from VRT paper Table VII, DAVIS. Only available for σ = 10, 20, 30, 40, 50."],
            ["FastDVDNet (published)", "Published PSNR from VRT paper Table VII, DAVIS. Sanity check for live results."],
            [""],
            ["Noise levels are specified by target noisy PSNR. σ = 255 / 10^(PSNR/20)."],
            ["Green-highlighted cells indicate the best performer for each noise level."],
        ]
        for r, row_data in enumerate(notes, 1):
            for c, val in enumerate(row_data, 1):
                ws_notes.cell(r, c, val)
                if r == 1:
                    ws_notes.cell(r, c).font = Font(bold=True, size=14, name="Arial", color="2F5496")
                elif c == 1 and r > 2:
                    ws_notes.cell(r, c).font = Font(bold=True, size=10, name="Arial")
                else:
                    ws_notes.cell(r, c).font = data_font

        ws_notes.column_dimensions['A'].width = 22
        ws_notes.column_dimensions['B'].width = 80

        wb.save(save_path)
        print(f"Excel report saved to {save_path}")

    def save_html_report(self, results, save_path):
        """Save benchmark results as a clean HTML report for visual inspection."""
        all_names = set()
        for range_data in results['noise_ranges'].values():
            all_names.update(range_data['denoisers'].keys())
        all_names = sorted(all_names)

        def sort_key(x):
            if x.startswith('PSNR='):
                return int(x.split('=')[1].replace('dB', ''))
            if x.startswith('σ='):
                return int(x.split('=')[1])
            return int(x.split('-')[0])

        sorted_ranges = sorted(
            results['noise_ranges'].keys(),
            key=sort_key,
            reverse=True
        )

        def make_table(metric, title, fmt, higher_better=True):
            html = f'<h2>{title}</h2>\n<table>\n<tr><th>Noise Level</th><th>σ</th><th>Noisy Baseline</th>'
            for n in all_names:
                html += f'<th>{n}</th>'
            html += '</tr>\n'

            for rk in sorted_ranges:
                data = results['noise_ranges'][rk]
                noisy_val = data['noisy'].get(metric)
                sigma = data.get('noise_std_used', 0)
                html += f'<tr><td class="range">{rk}</td><td>{sigma:.1f}</td><td>{noisy_val:{fmt}}</td>'

                # Find best
                vals = {}
                for name in all_names:
                    d = data['denoisers'].get(name, {})
                    v = d.get(metric)
                    if v is not None:
                        vals[name] = v
                best = max(vals, key=vals.get) if vals else None

                for name in all_names:
                    d = data['denoisers'].get(name, {})
                    v = d.get(metric)
                    if v is not None:
                        cls = ' class="best"' if name == best else ''
                        html += f'<td{cls}>{v:{fmt}}</td>'
                    else:
                        html += '<td class="na">N/A</td>'
                html += '</tr>\n'

            html += '</table>\n'
            return html

        html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>Denoiser Benchmark Results</title>
<style>
    body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 40px; background: #f8f9fa; color: #333; }}
    h1 {{ color: #2F5496; border-bottom: 3px solid #2F5496; padding-bottom: 10px; }}
    h2 {{ color: #2F5496; margin-top: 30px; }}
    table {{ border-collapse: collapse; width: 100%; margin: 15px 0; background: white; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
    th {{ background: #2F5496; color: white; padding: 10px 14px; text-align: center; font-size: 13px; }}
    td {{ padding: 8px 14px; text-align: center; border-bottom: 1px solid #e9ecef; font-size: 13px; }}
    tr:hover {{ background: #f1f3f5; }}
    .range {{ font-weight: bold; background: #e8edf3; }}
    .best {{ background: #c6efce; font-weight: bold; color: #006100; }}
    .na {{ color: #aaa; font-style: italic; }}
    .notes {{ background: white; padding: 20px; border-radius: 6px; margin-top: 30px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
    .notes p {{ margin: 6px 0; line-height: 1.6; }}
    .notes strong {{ color: #2F5496; }}
    .meta {{ color: #666; font-size: 12px; margin-top: 20px; }}
</style></head><body>
<h1>Video Denoiser Benchmark Results</h1>
<p class="meta">Generated: {results['metadata']['date']} | 
{results['metadata']['num_videos']} test videos | 
Seed: {results['metadata']['seed']} | 
Resolution: {results['metadata']['resize_to']}</p>
"""
        html += make_table('psnr', 'PSNR Comparison (dB) — Higher is Better', '.2f')
        html += make_table('ssim', 'SSIM Comparison — Higher is Better', '.4f')

        # PSNR gain table (custom since it's derived)
        html += '<h2>PSNR Improvement over Noisy Baseline (dB)</h2>\n<table>\n'
        html += '<tr><th>Noise Range (σ)</th>'
        for n in all_names:
            html += f'<th>{n}</th>'
        html += '</tr>\n'

        for rk in sorted_ranges:
            data = results['noise_ranges'][rk]
            noisy_p = data['noisy']['psnr']
            html += f'<tr><td class="range">{rk}</td>'

            gains = {}
            for name in all_names:
                d = data['denoisers'].get(name, {})
                p = d.get('psnr')
                if p is not None:
                    gains[name] = p - noisy_p
            best = max(gains, key=gains.get) if gains else None

            for name in all_names:
                d = data['denoisers'].get(name, {})
                p = d.get('psnr')
                if p is not None:
                    g = p - noisy_p
                    cls = ' class="best"' if name == best else ''
                    html += f'<td{cls}>+{g:.2f}</td>'
                else:
                    html += '<td class="na">N/A</td>'
            html += '</tr>\n'
        html += '</table>\n'

        html += """
<div class="notes">
<h2>Notes</h2>
<p><strong>Noise levels</strong> are specified by target noisy PSNR (dB). The corresponding σ is derived as: σ = 255 / 10^(PSNR/20).</p>
<p><strong>YourUNet:</strong> Blind denoiser trained with noise sampled uniformly in PSNR space for balanced perceptual coverage.</p>
<p><strong>FastDVDNet:</strong> Trained on σ ∈ [5, 55] and requires noise sigma as input. Expected to degrade significantly above σ ≈ 55.</p>
<p><strong>VRT (published):</strong> State-of-the-art transformer model. PSNR from Table VII of VRT paper (DAVIS testset). Matched to nearest PSNR level.</p>
<p><strong>FastDVDNet (published):</strong> Published PSNR from Table VII of VRT paper. Used as sanity check for live FastDVDNet results.</p>
<p><strong>Green cells</strong> indicate the best performer for each noise level.</p>
<p><strong>PSNR:</strong> Peak Signal-to-Noise Ratio — measures pixel-level accuracy in decibels. Higher is better.</p>
<p><strong>SSIM:</strong> Structural Similarity Index — measures perceptual quality (0 to 1). Higher is better.</p>
</div>
</body></html>"""

        with open(save_path, 'w') as f:
            f.write(html)
        print(f"HTML report saved to {save_path}")


# ============================================================================
#  Convenience function to run from notebook
# ============================================================================

def run_benchmark(
    your_model,
    davis_root,
    fastdvdnet_repo_path=None,
    device='cuda',
    max_videos=5,
    max_frames_per_video=10,
    resize_to=(480, 856),
    save_dir='./denoiser_evaluation'
):
    """
    Run the full benchmark from your training notebook.

    Args:
        your_model: Your trained BlindVideoDenoiserUNet (already loaded, eval mode)
        davis_root: Path to DAVIS dataset folder
        fastdvdnet_repo_path: Path to cloned FastDVDNet repo (e.g., '/content/fastdvdnet')
        device: 'cuda' or 'cpu'
        max_videos: Number of test videos
        max_frames_per_video: Frames per video
        resize_to: Processing resolution
        save_dir: Directory to save all outputs (JSON, Excel, HTML)

    Returns:
        results dict

    Usage in notebook:
        from benchmark import run_benchmark
        results = run_benchmark(
            your_model=model,
            davis_root='/content/DAVISDataset',
            fastdvdnet_repo_path='/content/fastdvdnet',
            save_dir='/content/drive/MyDrive/ResearchProject/denoiser_evaluation'
        )
    """
    save_dir = Path(save_dir)
    save_dir.mkdir(parents=True, exist_ok=True)

    results_json = save_dir / "results_data.json"
    results_xlsx = save_dir / "comparison_tables.xlsx"
    results_html = save_dir / "comparison_report.html"

    benchmark = DenoiserBenchmark(davis_root, device=device)

    # Register your model
    your_wrapper = YourDenoiserWrapper(your_model, device=device)
    benchmark.register_denoiser("YourUNet", your_wrapper)

    # Register FastDVDNet if repo available
    if fastdvdnet_repo_path and os.path.exists(fastdvdnet_repo_path):
        try:
            fdvd_wrapper = FastDVDNetWrapper(fastdvdnet_repo_path, device=device)
            benchmark.register_denoiser("FastDVDNet", fdvd_wrapper)
        except Exception as e:
            print(f"Warning: Could not load FastDVDNet: {e}")
            print("Continuing without FastDVDNet...")
    else:
        print(f"FastDVDNet repo not found at {fastdvdnet_repo_path}")
        print("To include FastDVDNet:")
        print("  !git clone https://github.com/m-tassano/fastdvdnet /content/fastdvdnet")

    # Run benchmark
    results = benchmark.run(
        max_videos=max_videos,
        max_frames_per_video=max_frames_per_video,
        resize_to=resize_to,
        results_path=str(results_json),
    )

    # Save Excel and HTML reports
    benchmark.save_excel_report(results, str(results_xlsx))
    benchmark.save_html_report(results, str(results_html))

    print(f"\n{'='*60}")
    print(f"All outputs saved to: {save_dir}")
    print(f"  - {results_json.name:30s} (raw data, accumulates history)")
    print(f"  - {results_xlsx.name:30s} (formatted spreadsheet)")
    print(f"  - {results_html.name:30s} (visual report)")
    print(f"{'='*60}")

    return results


if __name__ == "__main__":
    print("=== Denoiser Benchmark ===")
    print()
    print("Usage in your Colab notebook:")
    print()
    print("  # 1. Download FastDVDNet pretrained weights")
    print("  !git clone https://github.com/m-tassano/fastdvdnet /content/fastdvdnet")
    print()
    print("  # 2. Load your trained model")
    print("  from unet_denoiser import BlindVideoDenoiserUNet")
    print("  model = BlindVideoDenoiserUNet(in_channels=9, out_channels=3, base_channels=64, num_stages=3)")
    print("  ckpt = torch.load('/content/checkpoints/best_model.pt', map_location='cuda')")
    print("  model.load_state_dict(ckpt['model_state_dict'])")
    print("  model.eval()")
    print()
    print("  # 3. Run benchmark (saves JSON + Excel + HTML to the folder)")
    print("  from benchmark import run_benchmark")
    print("  results = run_benchmark(")
    print("      your_model=model,")
    print("      davis_root='/content/DAVISDataset',")
    print("      fastdvdnet_repo_path='/content/fastdvdnet',")
    print("      save_dir='/content/drive/MyDrive/ResearchProject/denoiser_evaluation'")
    print("  )")